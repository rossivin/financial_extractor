import os
import openpyxl as xl
import re

#Format is [account, date, description, debit, credit, balance, category, sub category, month, year]

this_folder = os.path.dirname(os.path.abspath(__file__))
pc_log = os.path.join(this_folder, 'report.csv')
pc_file = open(pc_log)
pc_info = pc_file.readlines()

xl_path = os.path.join(this_folder, 'export.xlsx')
xl_file = xl.Workbook()

sheet = xl_file['Sheet']

def extract_pc(r, info_file): #extract row 'r' information from PC Financial
    info = re.findall('"([^"]*)"', info_file[r])
    description = info[0]
    transaction_type = info[1]
    transaction_date = info[3]
    transaction_amount = -float(info[5])
    return["PC Financial Card", transaction_date, description, transaction_amount, transaction_type]

def print_pc(r, data_list, destination_sheet):
    destination_sheet['A' + str(r)].value = data_list[0]
    destination_sheet['B' + str(r)].value = data_list[1]
    destination_sheet['C' + str(r)].value = data_list[2]
    destination_sheet['D' + str(r)].value = data_list[3]

def pc_main(info_file, destination_sheet):
    print_row = destination_sheet.max_row + 1
    for i in range(len(info_file)-1,0,-1):
        pc_data = extract_pc(i, info_file)
        if pc_data[4] != "PAYMENT":
            print_pc(print_row, pc_data, destination_sheet)
            print_row += 1

def extract_td(r, info_file):
    info = info_file[r].strip().split(',')
    print(info)
    transaction_date = info[0]
    description = info[1]
    if info[2] != '':
        debit = float(info[2])
    else:
        debit = 0
    if info[3] != '':
        credit = float(info[3])
    else:
        credit = 0
    balance = float(info[4])
    return [transaction_date, description, debit, credit, balance]

def print_td(r, data_list, destination_sheet, account_name):
    destination_sheet['A' + str(r)].value = account_name
    destination_sheet['B' + str(r)].value = data_list[0]
    destination_sheet['C' + str(r)].value = data_list[1]
    destination_sheet['D' + str(r)].value = data_list[2]
    destination_sheet['E' + str(r)].value = data_list[3]
    destination_sheet['F' + str(r)].value = data_list[4]

def td_main(info_file, destination_sheet, account_name):
    print_row = destination_sheet.max_row + 1
    for i in range(1,len(info_file)):
        td_data = extract_td(i, info_file)
        if td_data[1] != "PAYMENT - THANK YOU":
            print_td(print_row, td_data, destination_sheet, account_name)
            print_row += 1

pc_log = os.path.join(this_folder, 'report.csv')
pc_file = open(pc_log)
pc_info = pc_file.readlines()
td_vcc_log = os.path.join(this_folder, 'cc-vinny.csv')
td_vcc_file = open(td_vcc_log)
td_vcc_info = td_vcc_file.readlines()
td_ccc_log = os.path.join(this_folder, 'cc-camila.csv')
td_ccc_file = open(td_ccc_log)
td_ccc_info = td_ccc_file.readlines()
td_joint_log = os.path.join(this_folder, 'joint.csv')
td_joint_file = open(td_joint_log)
td_joint_info = td_joint_file.readlines()

pc_main(pc_info, sheet)
td_main(td_vcc_info, sheet, "Vinny's Credit Card")
td_main(td_ccc_info, sheet, "Camila's Credit Card")
td_main(td_joint_info, sheet, "Joint Account")
xl_file.save(xl_path)
